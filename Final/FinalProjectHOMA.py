#!/usr/bin/env python3
import matplotlib.pyplot as plt
import mpld3;
from openpyxl import load_workbook
from scipy.signal import find_peaks



WorkBookFile = load_workbook('Book2.xlsx');
OutPut= open("OutPut.html", "w");
Alpha=2;

####### Determining the style and text of headers and paragraphs on the HTML file #########
OutPut.write("<html>\n")
OutPut.write("<head>\n")
OutPut.write("<style> h1 {color:hotpink;} body {background: lightgray; font-family:arial; text-align:justify;} </style> \n")
OutPut.write("<style> h2 {color:darkblue;} </style> \n")
OutPut.write("<style> h3 {color:green;} </style> \n")
OutPut.write("<h1>%s</h1> <h2>%s</h2> <h3>%s<h3>\n" % ("Programming Final Project\n", "Analyzing  Mass Spectrometry data of PEGylated Interferon beta-1b\n", "Created by: Homa Farahani"))
OutPut.write("<head>\n")
OutPut.write("<body>\n")
text="""My project has two components:<br></br>
1. It uses two excel sheets of Mass Spectrometry data and identifies all the significant peaks with an intensity threshold above 2. Data from each sheet is analyzed and plotted using an interactive interface. Selected numbers are shown in python environment as two distinct columns.<br></br>
2. It also reads Interferon beta-1b amino acid sequence, finds digestion positions at Lysin(K) and Arginine(R) sites, and predicts produced peptide fragments after cleavage with Trypsin enzyme.<br><br>Once sequences of produced peptides are identified, it could calculate the m/z ratio of each fragment up to three positive charges, then compare these numbers with Mass Spectrometry results to assess which one of these charged peptides exist in our data. The advantage of my code is that the user does not have to manually analyze Mass Spectrometry data, and could have rapid access to significant peaks, charge of the protein peptides, and could conveniently calculate the mass of each peptide.\n
Also, using a library called mpld3, a user could have a detailed view of all peaks, detected peptide peaks, and related intensity and m/z ratio of each fragment. Ability to scale the plot is crucial in evaluating the data."""
OutPut.write("<h4><p>%s</h4></p>\n"% text)
OutPut.write("</html>\n")

########## Defining function to plot the peak values including m/z and intensity on HTML file.########
######### The mpld3 library enables us to magnify and move the plot with appearing icons on plot's left bottom side ###########

def plotting(PeakMassPerZ,PeakIntensity,MassPerZ_value,pk_value,In,pk,figName,OutPut):
    plt.xlabel('MassPerZ (m/z)');
    plt.ylabel('Intensity');
    plt.title(figName);
    plt.plot(PeakMassPerZ,PeakIntensity,MassPerZ_value, pk_value,'g^',pk,In,'r^', markersize=3);
    Fig=plt.figure(num=1, figsize=(16, 20), dpi=250, facecolor='w', edgecolor='k')
    plt.savefig(figName+".png");
    Img=mpld3.fig_to_html(Fig);
    OutPut.write('<p>%s</p> \n' % (Img))
    #mpld3.show();
    #plt.show();
    plt.close(fig=None);


#####Defining a function to parse and read input excel file sheets cell by cell#########

def parse_set(Worksheet,sheetNumb):
    MassPerZset = [];
    intensitySet = [];
    Set=[MassPerZset,intensitySet];
    columnNumber=0;
    for column in Worksheet.iter_cols(0,2):
        if columnNumber>0:
            Set=[MassPerZset,intensitySet];
        rowNumber=0;
        for cell in column:         
            if cell.value is not None and columnNumber==0 and rowNumber>=1:
                MassPerZset.append(cell.value);
            if cell.value is not None and columnNumber==1 and rowNumber>=1:
                intensitySet.append(cell.value);
            rowNumber+=1;
        columnNumber+=1;
    return(Set);
    
############## Defining function to find peaks (intensity>2) in the excel file using Openpyxl library ################

def findPeak(InputDataSet,Alpha):
    intensity=InputDataSet[1];
    MassPerZ=InputDataSet[0];
    sum=0;
    for y in intensity:
        sum+=y
    BaseLine=Alpha*(sum/len(intensity));
    peaks, _ = find_peaks(intensity, height= BaseLine);
    pk_value = [];
    MassPerZ_value=[];
    line=("\n\n      MassPerZ   Intensity\n"); 
    print(line);
#   OutPut.write(line);
    for i in peaks:
        pk_value.append(intensity[i])
        MassPerZ_value.append(MassPerZ[i])
        line=("       %s      %s \n"%(MassPerZ[i],intensity[i]));
        print(line);
#       OutPut.write("<pre>" + line);
    return[pk_value,MassPerZ_value];

######### Defining a function that includes all amino acids' mass, so it could calculate the weight(mass) of produced peptides (after digestion with Trypsin) ############
def SubSeqWeight(sub):
    Residue=["G","A","S","P","V","T","C","L","I","N","D","Q","K","E","M","H","F","R","Y","W"];
    Weight=[57.021,71.037,87.032,97.053,99.068,101.048,103.009,113.084,113.084,114.043,115.027,128.059,128.095,129.043,131.04,137.059,147.068,156.101,163.063,186.079];
    SeqWeight=0%2;
    for i in sub:
        r=Residue.index(i);
        w=Weight[r];        
        SeqWeight+=w;
    return(SeqWeight);


########### Defining a function to assess the presence of each produced peptid through comparing m/z of them with ###########
########### all m/z peaks of MassSpectrometry data ##########
def findNearest_MassPerZ(PeakMassPerZ,weight):
    peak=0.0;
    for p in PeakMassPerZ:
        if (p-0.5) <= weight and p+0.5>=weight:
            peak=p;
            break;
    return(peak);


######### Parse and read each sheet of the excel file and find significant peaks with intensity above 2 (alpha)#########
k=0
PeakIntensity=[];
PeakMassPerZ=[];
InputDataSet=[];
for sheet in WorkBookFile:
    Set=parse_set(sheet,k);
    k+=1     
    MassPeak=findPeak(Set,Alpha);
    PeakIntensity.append(MassPeak[0]);
    PeakMassPerZ.append(MassPeak[1]);
    InputDataSet.append(Set);

###### Finding peptide fragments after enzyme digestion on "K" and "R" amino acid sites. In the following,########
######it calculates the m/z of these peptides for 1+, 2+, and 3+ charges, so it could check them later with closest peak number and find which of these charged fragments########
######are found among MassSpect resulted peaks########

seq="SYNLLGFLQRSSNFQSQKLLWQLNGRLEYCLKDRMNFDIPEEIKQLQQFQKEDAALTIYEMLQNIFAIFRQDSSSTGWNETIVENLLANVYHQINHLKTVLEEKLEKEDFTRGKLMSSLHLKRYYGRILHYLKAKEYSHCAWTIVRVEILRNFYFINRLTGYLRN";
DigestPos=[];
for i in range (len(seq)):
    if seq[i]=="R" or seq[i]=="K":
        DigestPos.append(i);
DigestPos.append(len(seq)-1)
line=("Trypsin digestion positions on the given sequence, and expected peptide fragments of Interferon beta-1b: \n %s" %str(DigestPos));
print(line);
OutPut.write("<h4> Trypsin digestion positions on the given sequence, and expected peptide fragments of Interferon beta-1b: <br><br> %s </h4>" %str(DigestPos));
OutPut.write("<img src='Picture1.png' height=550; style = \"position:relative; margin-right:250px; top:120px; float:right;\"/>")


subseq=[];
subWeight=[];
PK_0=[];
In_0=[];
PK_1=[];
In_1=[];
start=0;
for i in DigestPos:
    end=i+1;       
    sub=seq[start:end];
    subseq.append(sub);
    start=end;
for sub in subseq:
    line=("\nPeptide subsequence:  %s \n\n Possible m/z ratio of this predicted fragment                    m/z (+),     m/z (2+),    m/z (3+)"%(sub));
    print(line);
    #OutPut.write("<pre>" +line + "</pre> <br>");
    OutPut.write("<h5><li> %s </li></h5>" %sub);
    
    weight=SubSeqWeight(sub);
    weight=weight+18;
    weight=round(weight, 2)
    weight1=(weight+1)
    weight2=(weight+2)/2
    weight3=(weight+3)/3
    weight3=round(weight3, 2)
    subWeight.append(weight);
    peak1=findNearest_MassPerZ(PeakMassPerZ[0],weight1)
    peak2=findNearest_MassPerZ(PeakMassPerZ[0],weight2)
    peak3=findNearest_MassPerZ(PeakMassPerZ[0],weight3)
    peak11=findNearest_MassPerZ(PeakMassPerZ[1],weight1)
    peak12=findNearest_MassPerZ(PeakMassPerZ[1],weight2)
    peak13=findNearest_MassPerZ(PeakMassPerZ[1],weight3)
    peaks_0=[peak1,peak2,peak3];
    peaks_1=[peak11,peak12,peak13];
    weights=[weight1,weight2,weight3];
    line=("\n Mass per charge(z) for 3 different charges: %29.4f,%12.4f,%12.4f"%(weight1,weight2,weight3));
    print(line);  
    #OutPut.write("<pre>" +line + "</pre> <br>");
    line=("\n Coressponding Peak Intinsity in the data on the first sheet: %12.4f,%12.4f,%12.4f"%(peak1,peak2,peak3));
    print(line);
    #OutPut.write("<pre>" +line + "</pre> <br>");
    line=("\n Coressponding Peak Intinsity in the data on the second sheet: %11.4f,%12.4f,%12.4f \n"%(peak11,peak12,peak13));
    print(line);
    #OutPut.write("<pre>" +line + "</pre> <br>");

########## Make a list of found peptides and write their m/z ratio ###########
    PK=[];
    In=[]
    for p in peaks_0:
        if p>0.0:
            PK_0.append(p);
            j=InputDataSet[0][0].index(p);
            inten=InputDataSet[0][1][j];
            In_0.append(inten);
    In.append(In_0)
    PK.append(PK_0);  
    for p in peaks_1:
        if p>0.0:
            PK_1.append(p);
            j=InputDataSet[1][0].index(p);
            inten=InputDataSet[1][1][j];
            In_1.append(inten);
    In.append(In_1)
    PK.append(PK_1);  

line=("\n\n m/z ratio of found peptides (among three calculated charged variants) on sheet one:\n");
print(line);
OutPut.write("<hr>" +line + "</hr>");
line=(str(PK_0));
print(line);
OutPut.write(line + "<br><br>");

line=("\n\n m/z ratio of found peptides (among three calculated charged variants) on sheet two:\n");
print(line);
OutPut.write(line + "</pre>");
line=(str(PK_1));
print(line);
OutPut.write(line + "</pre>");
OutPut.write("<hr></hr>");


########### Generating the HTML file, and plots of significant peaks and Interferon beta-1b peptide fragments for each sheet #########

text='''In the following, plots are showing significant peaks (green markers) of Mass Spectrometry data (each plot represents peaks of related data sheet). Red markers are indicating peaks related to Interferon beta-1b peptides found among all Mass Spectrometry peaks.<br></br>
*** Icons appear at the left bottom of each plot enable the user to magnify and transfer plot peaks.***'''
OutPut.write("<h4><p>%s</h4></p>" %text)

for i in range (2):
	plotting(InputDataSet[i][0],InputDataSet[i][1],PeakMassPerZ[i],PeakIntensity[i],In[i],PK[i],"Mass Spectrometry Peaks Plot_%d"%i,OutPut);

OutPut.close()